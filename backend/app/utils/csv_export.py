import csv
import io
from fastapi.responses import StreamingResponse
import openpyxl
from typing import Dict, List, Any

def create_csv_response(data: List[Dict[str, Any]], headers: List[str], filename: str) -> StreamingResponse:
    """Create a CSV streaming response from data"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    
    for row in data:
        writer.writerow([row.get(header, "") for header in headers])
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def create_csv_template(headers: List[str], sample_data: List[str], filename: str) -> StreamingResponse:
    """Create a CSV template with headers and sample data"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample_data)
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def create_excel_response(worksheets: Dict[str, Dict], filename: str) -> StreamingResponse:
    workbook = openpyxl.Workbook()
    
    workbook.remove(workbook.active)
    
    for sheet_name, sheet_data in worksheets.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        headers = sheet_data['headers']
        data = sheet_data['data']
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
